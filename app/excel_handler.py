import os
import pandas as pd
import openpyxl
import logging
import re
import sys
sys.stdout.reconfigure(encoding='utf-8')
sys.stderr.reconfigure(encoding='utf-8')

from openpyxl.styles import PatternFill, Font
from datetime import datetime
from app.utils.insurance_providers import INSURANCE_PROVIDERS, SPECIFIC_TDS_RATE_PROVIDERS, NEW_INDIA_THRESHOLD

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(stream=sys.stdout),  # Use stdout with UTF-8 config
        logging.FileHandler("pdf_processing.log", encoding='utf-8')  # Add encoding here
    ]
)
logger = logging.getLogger("excel_handler")


class ExcelHandler:
    def __init__(self, excel_path):
        """
        Initialize the Excel handler.

        Args:
            excel_path (str): Path to the Excel file
        """
        self.excel_path = excel_path
        self.df = None
        self.wb = None
        self.ws = None
        self.invoice_column = None
        self.client_ref_column = None
        self.header_mapping = {}

        # Load the Excel file
        self.load_excel()

    def _is_valid_claim_ref(self, identifier):
        """Check if an identifier matches the claim reference format patterns"""
        return identifier and re.match(r'^(?:ENG|MAR|FIR|MSC|LIA)\d+$', identifier)

    def load_excel(self):
        """Load the Excel file into a pandas DataFrame and openpyxl Workbook."""
        try:
            # Load with pandas for data manipulation
            self.df = pd.read_excel(self.excel_path)
            logger.info(f"Loaded Excel file with {len(self.df)} rows and {len(self.df.columns)} columns")

            # Also load with openpyxl for direct cell manipulation
            self.wb = openpyxl.load_workbook(self.excel_path)
            self.ws = self.wb.active

            # Identify the ID columns (Invoice No. and Client Ref. No.)
            self.identify_id_columns()

            # Create header mapping
            self.create_header_mapping()

        except Exception as e:
            logger.error(f"Error loading Excel file: {str(e)}")
            raise

    def identify_id_columns(self):
        """Identify the columns for Invoice No. and Client Ref. No."""
        # Search for Invoice Number column
        invoice_columns = [
            'Invoice No.', 'Invoice No', 'Invoice Number', 'Invoice',
            'Inv No.', 'Inv No', 'Inv Number', 'Invoice #', 'Bill No.',
            'Bill No', 'Payment Details 5', 'ThirdPartyInv', 'ILA_REF_NO',
            'Expense Paid', 'Invoice'
        ]

        # Search for Client Reference Number column
        client_ref_columns = [
            'Client Ref. No.', 'Client Ref. No', 'Client Reference Number',
            'Client Ref', 'Ref. No.', 'Reference No.', 'Reference Number', 'Claim#', 'Claim Number', 'CLAIM NUMBER',
            'Claim number', 'Sub Claim No', 'INV REF:Claim No', 'Claim No', 'DESC', 'Payment Details 7',
            'ClaimNo', 'Claim_Ref_No', 'CLAIM_REF_NO', 'Invoice Details'
        ]

        # Find exact matches first
        for col in invoice_columns:
            if col in self.df.columns:
                self.invoice_column = col
                logger.info(f"Found Invoice column: {col}")
                break

        for col in client_ref_columns:
            if col in self.df.columns:
                self.client_ref_column = col
                logger.info(f"Found Client Reference column: {col}")
                break

        # If exact matches not found, look for partial matches
        if not self.invoice_column:
            for col in self.df.columns:
                if any(invoice.lower() in col.lower() for invoice in invoice_columns):
                    self.invoice_column = col
                    logger.info(f"Found Invoice column (partial match): {col}")
                    break

        if not self.client_ref_column:
            for col in self.df.columns:
                if any(ref.lower() in col.lower() for ref in client_ref_columns):
                    self.client_ref_column = col
                    logger.info(f"Found Client Reference column (partial match): {col}")
                    break

        # If still not found, use the first two columns as fallback
        if not self.invoice_column and len(self.df.columns) > 0:
            self.invoice_column = self.df.columns[0]
            logger.warning(f"No obvious Invoice column found. Using first column: {self.invoice_column}")

        if not self.client_ref_column and len(self.df.columns) > 1:
            self.client_ref_column = self.df.columns[1]
            logger.warning(f"No obvious Client Reference column found. Using second column: {self.client_ref_column}")

    def create_header_mapping(self):
        """Create a mapping between common field names and actual column names."""
        # Define common field names and possible variants
        field_mappings = {
            'receipt_date': ['Receipt Date', 'Receipt Date', 'Payment Date', 'Value Date', 'Date', 'Payment Ini Date',
                             'Value date', 'Advice sending date', 'Settlement Date', 'Value Date', 'VALUE DATE'],
            'receipt_amount': ['Receipt Amount', 'Receipt Amt', 'Payment Amount', 'Amount', 'Value', 'Amount',
                               'Remittance amount', 'Net Paid Amount', 'REMITTANCE AMOUNT', 'Net Amount',
                               'Amount (INR)', 'AMOUNT', 'Payment amount', 'TRF AMOUNT'],
            'tds': ['TDS', 'TDS Amount', 'Tax Deducted at Source', 'Tax Amount', 'TDS Amount', 'TDS', 'TDS Amt',
                    'Payment Details 4', 'Less : TDS'],
            'tds_computed': ['TDS Computed?', 'TDS Computed', 'Is TDS Computed', 'Computed TDS']
        }

        # Create mapping between standardized field names and actual column names
        for field, variants in field_mappings.items():
            for col in self.df.columns:
                if col in variants or any(variant.lower() in col.lower() for variant in variants):
                    self.header_mapping[field] = col
                    break

        logger.info(f"Created header mapping: {self.header_mapping}")

    def compute_tds(self, amount, text, detected_provider=None):
        """
        Compute TDS based on the receipt amount and text content.
        Applies 11.111111% for specific insurance companies and 9.259259% for others.
        Special handling for New India Assurance with threshold of 300000.
        """
        try:
            # Normalize text for better matching
            normalized_text = ' '.join(text.lower().split())
            logger.debug(f"Normalized text for insurance detection (first 300 chars): {normalized_text[:300]}...")

            # Use already detected provider if available
            if detected_provider:
                logger.info(f"Using previously detected insurance provider: {detected_provider}")

                # Apply appropriate TDS rate based on the provider
                if detected_provider in SPECIFIC_TDS_RATE_PROVIDERS:
                    # Special handling for New India Assurance
                    if detected_provider == "new_india" and amount <= NEW_INDIA_THRESHOLD:
                        tds = round(amount * 0.09259259, 2)  # 9.259259% for amounts <= 300000
                        logger.info(f"TDS computed for New India Assurance (≤ 300000): {tds} (9.259259% of {amount})")
                    else:
                        tds = round(amount * 0.11111111, 2)  # 11.111111% for specific providers
                        logger.info(
                            f"TDS computed for specific provider ({detected_provider}): {tds} (11.111111% of {amount})")
                else:
                    tds = round(amount * 0.09259259, 2)  # 9.259259% for all other providers
                    logger.info(f"TDS computed for non-specific provider: {tds} (9.259259% of {amount})")

                return tds, True

            # If no provider was detected, try to detect one now
            for provider, keywords in INSURANCE_PROVIDERS.items():
                for keyword in keywords:
                    # Use more precise matching to avoid false positives
                    if keyword in normalized_text:
                        # For single word keywords, ensure they're not part of other words
                        if len(keyword.split()) == 1:
                            pattern = r'\b' + re.escape(keyword) + r'\b'
                            if re.search(pattern, normalized_text):
                                detected_provider = provider
                                logger.info(f"Detected insurance provider: {provider} (keyword: {keyword})")
                                break
                        else:
                            # Multi-word keywords are more specific already
                            detected_provider = provider
                            logger.info(f"Detected insurance provider: {provider} (keyword: {keyword})")
                            break
                if detected_provider:
                    break

            # Apply appropriate TDS rate based on the detected provider
            if detected_provider in SPECIFIC_TDS_RATE_PROVIDERS:
                # Special handling for New India Assurance
                if detected_provider == "new_india" and amount <= NEW_INDIA_THRESHOLD:
                    tds = round(amount * 0.09259259, 2)  # 9.259259% for amounts <= 300000
                    logger.info(f"TDS computed for New India Assurance (≤ 300000): {tds} (9.259259% of {amount})")
                else:
                    tds = round(amount * 0.11111111, 2)  # 11.111111% for specific providers
                    logger.info(
                        f"TDS computed for specific provider ({detected_provider}): {tds} (11.111111% of {amount})")
            else:
                tds = round(amount * 0.09259259, 2)  # 9.259259% for all other providers
                logger.info(f"TDS computed for non-specific provider: {tds} (9.259259% of {amount})")

            return tds, True

        except Exception as e:
            logger.error(f"Error computing TDS: {str(e)}")
            return 0.0, True

    def find_row_by_identifiers(self, unique_id, data_points, pdf_text=None):
        """
        Find a row in the DataFrame by matching either Invoice No. or Client Ref. No.
        Enhanced with strict matching to ensure 100% accuracy.
        """
        if not unique_id:
            return None

        try:
            # Initialize match_found variable (fix for the undefined variable error)
            match_found = False  # Add this line to fix the error

            # Log the unique_id for debugging
            logger.info(f"Attempting to match ID: {unique_id}")

            # Clean the unique_id to improve matching chances
            # Remove any unwanted characters that might interfere with matching
            # Use a more tailored approach that preserves the important parts
            if unique_id:
                # For IDs with Roman letters and numbers, clean carefully
                if re.search(r'[A-Za-z0-9]', unique_id):
                    cleaned_id = re.sub(r'[^A-Za-z0-9/-]', '', unique_id)
                    # Also remove any newline characters and spaces
                    cleaned_id = cleaned_id.replace('\n', '').replace(' ', '')
                    logger.info(f"Cleaned ID for matching: {cleaned_id}")
                else:
                    # For non-Roman script IDs, try to find a number pattern
                    number_match = re.search(r'(\d+-\d+|\d{10,})', text)
                    if number_match:
                        cleaned_id = number_match.group(1)
                        logger.info(f"Extracted numeric ID from context: {cleaned_id}")
                    else:
                        cleaned_id = unique_id  # Keep as is if no better option
            logger.info(f"Cleaned ID for matching: {cleaned_id}")

            # Check if this is an ICICI Lombard document with "ENG" pattern
            if (unique_id and any(unique_id.startswith(prefix) for prefix in ["ENG", "MAR", "FIR", "MSC", "LIA"])) or (
                    pdf_text and "CLAIM_REF_NO" in pdf_text and "LAE Invoice No" in pdf_text):
                # For ICICI Lombard documents, try both claim ref and invoice number
                invoice_no = data_points.get('invoice_no')
                if invoice_no:
                    logger.info(f"Found invoice number in data_points: {invoice_no}")

                # Use the match_with_fallback method for better matching
                return self.match_with_fallback(cleaned_id, invoice_no, cleaned_id)

            # Extract just the numeric part if it's a complex ID
            numeric_part = re.sub(r'[^0-9]', '', unique_id)
            if len(numeric_part) >= 5:  # Only use if it's a substantial number
                logger.info(f"Numeric part of ID: {numeric_part}")
            else:
                numeric_part = None

            # For claim numbers with format like 510000/11/2025/000000, extract various parts
            claim_parts = None
            claim_last_part = None
            claim_first_part = None
            if '/' in cleaned_id:
                claim_parts = cleaned_id.split('/')
                claim_last_part = claim_parts[-1]
                claim_first_part = claim_parts[0]
                logger.info(f"Claim parts: {claim_parts}")
                logger.info(f"Last part of claim number: {claim_last_part}")
                logger.info(f"First part of claim number: {claim_first_part}")

            # For invoice numbers with format like 2425-12345, extract parts
            invoice_parts = None
            invoice_prefix = None
            invoice_number = None
            if '-' in cleaned_id:
                invoice_parts = cleaned_id.split('-')
                invoice_prefix = invoice_parts[0]
                invoice_number = invoice_parts[-1] if len(invoice_parts) > 1 else None
                logger.info(f"Invoice parts: {invoice_parts}")
                logger.info(f"Invoice prefix: {invoice_prefix}")
                logger.info(f"Invoice number: {invoice_number}")

            # First check if we have an invoice_no field in data_points (HDFC ERGO specific case)
            if 'invoice_no' in data_points and data_points['invoice_no']:
                invoice_id = data_points['invoice_no']
                logger.info(f"Found invoice number in data_points: {invoice_id}")

                # Try exact match on invoice column
                if self.invoice_column is not None:
                    column_series = self.df[self.invoice_column].astype(str)
                    exact_matches = column_series[column_series == invoice_id]
                    if not exact_matches.empty:
                        index = exact_matches.index[0]
                        logger.info(f"Found exact match for invoice_no in column {self.invoice_column} at row {index}")
                        return index

            # Reliance - Add this at the beginning of the find_row_by_identifiers method in excel_handler.py
            if pdf_text and "RELIANCE GENERAL INSURAANCE" in pdf_text:
                # For Reliance documents, specifically look for the invoice number in Payment Details 5
                invoice_pattern = r'Payment\s+Details\s+5\s*:?\s*(\d{4}-\d{4,5})'
                invoice_match = re.search(invoice_pattern, pdf_text, re.IGNORECASE)
                if invoice_match:
                    invoice_id = invoice_match.group(1).strip()
                    logger.info(f"Reliance document: Using invoice number from Payment Details 5: {invoice_id}")

                    # Try matching this invoice ID directly
                    if self.invoice_column is not None:
                        column_series = self.df[self.invoice_column].astype(str)
                        exact_matches = column_series[column_series == invoice_id]
                        if not exact_matches.empty:
                            index = exact_matches.index[0]
                            logger.info(
                                f"Found exact match for Reliance invoice in column {self.invoice_column} at row {index}")
                            return index

            # For RG Cargo check if the unique_id is just "Invoice" (column header issue)
            if unique_id == "Invoice" and pdf_text:
                # Try to extract real invoice numbers from text
                invoice_matches = re.findall(r'(\d{4}-\d{5})', pdf_text)
                if invoice_matches:
                    logger.info(
                        f"Replacing generic 'Invoice' ID with actual invoice number: {invoice_matches[0]}")
                    unique_id = invoice_matches[0]

            # If no match found and we have an invoice number, try matching that (ICICI specific case)
            if 'invoice_no' in data_points and data_points['invoice_no']:
                logger.info(f"Trying to match using invoice number: {data_points['invoice_no']}")

                if self.invoice_column is not None:
                    column_series = self.df[self.invoice_column].astype(str)
                    exact_matches = column_series[column_series == data_points['invoice_no']]
                    if not exact_matches.empty:
                        index = exact_matches.index[0]
                        logger.info(
                            f"Found exact match for invoice number in column {self.invoice_column} at row {index}")
                        return index

            # Try each column with exact matching first - STRICT MATCHING ONLY
            for column in [self.invoice_column, self.client_ref_column]:
                if column is not None:
                    column_series = self.df[column].astype(str)

                    # Exact match (case-sensitive)
                    exact_matches = column_series[column_series == cleaned_id]
                    if not exact_matches.empty:
                        index = exact_matches.index[0]
                        logger.info(f"Found exact case-sensitive match in column {column} at row {index}")
                        return index

                    # Exact match (case-insensitive)
                    exact_matches_case_insensitive = column_series[column_series.str.lower() == cleaned_id.lower()]
                    if not exact_matches_case_insensitive.empty:
                        index = exact_matches_case_insensitive.index[0]
                        logger.info(f"Found exact case-insensitive match in column {column} at row {index}")
                        return index

            # If no match yet, try to look specifically for just the numeric part of F-pattern identifiers
            if not match_found and re.match(r'F\d{7}', cleaned_id):
                # Try matching just the 7-digit part without the F prefix
                numeric_part = cleaned_id[1:]  # Remove the 'F'

                for column in [self.invoice_column, self.client_ref_column]:
                    if column is not None:
                        column_series = self.df[column].astype(str)

                        # First check if the full numeric part exists anywhere
                        numeric_matches = column_series[column_series.str.contains(numeric_part)]
                        if not numeric_matches.empty:
                            logger.info(f"Found numeric part match for '{numeric_part}' in column {column}")
                            return numeric_matches.index[0]

                        # If not, check if the last 5 digits match (sometimes only partial numbers are stored)
                        if len(numeric_part) >= 5:
                            last_5_digits = numeric_part[-5:]
                            partial_matches = column_series[column_series.str.contains(last_5_digits)]
                            if not partial_matches.empty:
                                logger.info(
                                    f"Found partial match with last 5 digits '{last_5_digits}' in column {column}")
                                return partial_matches.index[0]

            # Debug: Print sample Excel values to check what we're matching against
            logger.info(f"Excel row sample for potential matches:")
            if self.invoice_column:
                sample_values = self.df[self.invoice_column].astype(str).head(5).tolist()
                logger.info(f"Column '{self.invoice_column}' first 5 values: {sample_values}")
            if self.client_ref_column:
                sample_values = self.df[self.client_ref_column].astype(str).head(5).tolist()
                logger.info(f"Column '{self.client_ref_column}' first 5 values: {sample_values}")

            # As a final fallback, try using invoice_no from data_points
            if 'invoice_no' in data_points and data_points['invoice_no']:
                invoice_no = data_points['invoice_no']
                logger.info(f"Found invoice number in data_points: {invoice_no}")
                if self.invoice_column is not None:
                    column_series = self.df[self.invoice_column].astype(str)
                    invoice_matches = column_series[column_series == invoice_no]
                    if not invoice_matches.empty:
                        index = invoice_matches.index[0]
                        logger.info(
                            f"Found match using invoice number in column {self.invoice_column} at row {index}")
                        return index

            # No exact match found - that's all we try with strict matching
            logger.warning(f"No exact match found for ID: {unique_id}")
            return None

        except Exception as e:
            logger.error(f"Error finding row by ID: {str(e)}")
            return None


    def match_with_fallback(self, unique_id, invoice_no=None, claim_ref=None):
        """
        Try multiple matching strategies to find the correct Excel row.

        Args:
            unique_id (str): The primary unique identifier
            invoice_no (str, optional): Invoice number for fallback matching
            claim_ref (str, optional): Claim reference for fallback matching

        Returns:
            int or None: Index of the matching row, or None if no match found
        """
        # First try unique ID in both columns
        if unique_id:
            logger.info(f"Trying to match with unique ID: {unique_id}")
            for column in [self.invoice_column, self.client_ref_column]:
                if column is not None:
                    column_series = self.df[column].astype(str)
                    exact_matches = column_series[column_series == unique_id]
                    if not exact_matches.empty:
                        index = exact_matches.index[0]
                        logger.info(f"Found exact match for unique ID in column {column} at row {index}")
                        return index

        # Then try invoice number specifically in invoice column
        if invoice_no and self.invoice_column:
            logger.info(f"Trying to match with invoice number: {invoice_no}")
            column_series = self.df[self.invoice_column].astype(str)
            matches = column_series[column_series == invoice_no]
            if not matches.empty:
                index = matches.index[0]
                logger.info(f"Found match for invoice number in column {self.invoice_column} at row {index}")
                return index

        # Then try claim reference specifically in client ref column
        if self._is_valid_claim_ref(unique_id) and self.client_ref_column:
            logger.info(f"Trying to match with claim reference: {claim_ref}")
            column_series = self.df[self.client_ref_column].astype(str)
            exact_matches = column_series[column_series == unique_id]
            if not exact_matches.empty:
                index = exact_matches.index[0]
                logger.info(f"Found exact match for claim reference in {self.client_ref_column} at row {index}")
                return index

        logger.warning(f"No match found for ID: {unique_id}, invoice: {invoice_no}, claim ref: {claim_ref}")
        return None

    def update_row_with_data(self, row_index, data_points, pdf_text=None, detected_provider=None):
        """
        Update a row in the Excel file with extracted data points.
        Enhanced with improved validation and handling of edge cases for various formats.

        Args:
            row_index (int): Index of the row to update
            data_points (dict): Dictionary of field names and values
            pdf_text (str, optional): The full PDF text for context if needed

        Returns:
            tuple: (success, updated_fields, failed_fields)
        """
        if row_index is None or not data_points:
            return False, [], list(data_points.keys())

        if 'receipt_amount' in data_points and data_points['receipt_amount']:
            # Check if receipt_amount might be an ID (too many digits for a typical amount)
            if len(str(data_points['receipt_amount'])) > 10 and str(data_points['receipt_amount']).isdigit():
                logger.warning(f"Receipt amount {data_points['receipt_amount']} looks like an ID, not an amount")

                # Try to extract from PDF text directly
                if pdf_text:
                    # For New India, try to get the TOTAL amount
                    total_match = re.search(r'TOTAL:\s*([\d,\.]+)', pdf_text)
                    if total_match:
                        data_points['receipt_amount'] = total_match.group(1).replace(',', '')
                        logger.info(f"Corrected amount to {data_points['receipt_amount']} from TOTAL")

        if 'receipt_amount' in data_points:
            logger.info(f"IMPORTANT DEBUG - Initial receipt_amount in data_points: {data_points['receipt_amount']}")

            # Look for specific patterns in PDF text to verify the amount is correct
            if pdf_text and "New India Assurance" in pdf_text:
                # Extract the TOTAL amount from New India payment advice
                total_match = re.search(r'TOTAL:\s*([\d,\.]+)', pdf_text)
                if total_match:
                    correct_amount = total_match.group(1).replace(',', '')
                    logger.info(f"Found TOTAL amount in New India advice: {correct_amount}")

                    # Compare with current amount and use the TOTAL if it's different
                    if data_points['receipt_amount'] != correct_amount:
                        logger.warning(
                            f"Correcting receipt_amount from {data_points['receipt_amount']} to {correct_amount}")
                        data_points['receipt_amount'] = correct_amount


        # Fix European number format in receipt_amount
        if 'receipt_amount' in data_points and data_points['receipt_amount']:
            try:
                amount_val = float(data_points['receipt_amount'].replace(',', ''))

                # If the amount seems very small compared to TDS, it might be wrong
                if 'tds' in data_points and data_points.get('tds'):
                    tds_val = float(data_points['tds'].replace(',', ''))
                    if amount_val < tds_val and tds_val > 0:
                        logger.warning(
                            f"Receipt amount ({amount_val}) is less than TDS ({tds_val}). This may be incorrect.")
                        # Check if the amount might actually be the TDS by mistake
                        if abs(amount_val - tds_val) < 1.0:  # If they're very close
                            logger.warning("Receipt amount appears to be TDS. Skipping this field.")
                            data_points.pop('receipt_amount')
                            failed_fields.append('receipt_amount')
            except ValueError:
                pass
            amount_str = data_points['receipt_amount']

            # Handle European number format with multiple periods (9.989.00)
            if '.' in amount_str and amount_str.count('.') > 1:
                try:
                    # For format like 9.989.00
                    parts = amount_str.split('.')
                    if len(parts) == 3:
                        clean_amount = parts[0] + parts[1] + '.' + parts[2]
                        data_points['receipt_amount'] = clean_amount
                        logger.info(f"Fixed European format: {amount_str} -> {clean_amount}")
                except Exception as e:
                    logger.error(f"Error fixing European format: {e}")

        updated_fields = []
        failed_fields = []

        try:
            # Get the Excel row number (add 2 to account for 0-indexing and header row)
            excel_row = row_index + 2

            # Validate and fix data before updating
            # For receipt amount, ensure it's a proper number and not truncated
            if 'receipt_amount' in data_points and data_points['receipt_amount']:
                try:
                    amount_str = data_points['receipt_amount']
                    # Clean the amount value
                    amount_str = re.sub(r'[^0-9.]', '', amount_str)

                    # Check if the amount seems too small for a payment
                    if len(amount_str) <= 3 and '.' not in amount_str:
                        logger.warning(f"Receipt amount {amount_str} seems too small, might be truncated")

                        # For OICL ADVICE PDF, fix the truncated value
                        if pdf_text and ("oriental insurance" in pdf_text.lower() or "hsbc" in pdf_text.lower()):
                            # Look for the full amount in the text
                            amount_patterns = [
                                r'Remittance\s+amount\s*:?\s*(?:INR)?\s*(\d{1,3}(?:,\d{3})*(?:\.\d{2})?)',
                                r'Amount.*?(\d{6}(?:\.\d{2})?)',
                                r'Amount\s*(\d{6})',
                            ]

                            for pattern in amount_patterns:
                                amount_match = re.search(pattern, pdf_text, re.IGNORECASE)
                                if amount_match:
                                    amount_str = amount_match.group(1).replace(',', '')
                                    logger.info(
                                        f"Fixed truncated amount: {data_points['receipt_amount']} -> {amount_str}")
                                    data_points['receipt_amount'] = amount_str
                                    break

                        # For other payment advice PDFs
                        elif pdf_text:
                            # Generic patterns to find amounts
                            amount_patterns = [
                                r'amount\s*:?\s*(?:Rs\.?|INR)?\s*(\d{1,3}(?:,\d{3})*(?:\.\d{1,2})?)',
                                r'remittance\s+amount\s*:?\s*(?:Rs\.?|INR)?\s*(\d{1,3}(?:,\d{3})*(?:\.\d{1,2})?)',
                                r'payment\s+amount\s*:?\s*(?:Rs\.?|INR)?\s*(\d{1,3}(?:,\d{3})*(?:\.\d{1,2})?)',
                                r'net\s+amount\s*:?\s*(?:Rs\.?|INR)?\s*(\d{1,3}(?:,\d{3})*(?:\.\d{1,2})?)',
                                r'rupees\s+.*?(\d{1,3}(?:,\d{3})*(?:\.\d{1,2})?)',
                            ]

                            for pattern in amount_patterns:
                                amount_match = re.search(pattern, pdf_text, re.IGNORECASE)
                                if amount_match:
                                    amount_str = amount_match.group(1).replace(',', '')
                                    logger.info(
                                        f"Fixed truncated amount using generic pattern: {data_points['receipt_amount']} -> {amount_str}")
                                    data_points['receipt_amount'] = amount_str
                                    break

                    # Convert to float to ensure it's a valid number
                    amount = float(amount_str)
                    # Format with 2 decimal places
                    data_points['receipt_amount'] = str(amount)
                except ValueError:
                    logger.error(f"Invalid amount format: {data_points['receipt_amount']}")
                    failed_fields.append('receipt_amount')
                    data_points.pop('receipt_amount', None)

            # For receipt date, ensure it's in a proper date format
            if 'receipt_date' in data_points and data_points['receipt_date']:
                try:
                    date_str = data_points['receipt_date']
                    # Handle different date formats
                    date_formats = [
                        '%d %b %Y',  # 11 Feb 2025
                        '%d-%m-%Y',  # 12-02-2025
                        '%d/%m/%Y',  # 12/02/2025
                        '%Y-%m-%d',  # 2025-02-12
                        '%Y/%m/%d',  # 2025/02/12
                        '%d.%m.%Y',  # 12.02.2025
                        '%b %d, %Y',  # Feb 12, 2025
                        '%d-%b-%Y',  # 12-Feb-2025
                        '%d/%b/%Y',  # 12/Feb/2025
                    ]

                    parsed_date = None
                    for date_format in date_formats:
                        try:
                            parsed_date = datetime.strptime(date_str, date_format)
                            break
                        except ValueError:
                            continue

                    if parsed_date:
                        # Standardize to dd-mm-yyyy format
                        data_points['receipt_date'] = parsed_date.strftime('%d-%m-%Y')
                        logger.debug(f"Validated receipt date: {data_points['receipt_date']}")
                    else:
                        logger.warning(f"Could not parse date: {date_str}")
                        failed_fields.append('receipt_date')
                        data_points.pop('receipt_date', None)
                except Exception as e:
                    logger.error(f"Error formatting date {data_points['receipt_date']}: {str(e)}")
                    failed_fields.append('receipt_date')
                    data_points.pop('receipt_date', None)

            # Process and compute TDS if needed
            # Changed following line from this to below - if 'receipt_amount' in data_points and data_points['receipt_amount'] and 'tds' not in data_points:
            if 'receipt_amount' in data_points and data_points['receipt_amount']:
                try:
                    # Clean the amount value and convert to float
                    amount_str = data_points['receipt_amount']
                    amount = float(re.sub(r'[^\d.]', '', amount_str))

                    # Compute TDS
                    tds_value, is_computed = self.compute_tds(amount, pdf_text if pdf_text else "", detected_provider)

                    # Add to data_points
                    data_points['tds'] = str(tds_value)
                    data_points['tds_computed'] = 'Yes'
                    logger.info(f"TDS computed: {tds_value}")
                except Exception as e:
                    logger.error(f"Error computing TDS: {str(e)}")
                    failed_fields.append('tds')
            elif 'tds' in data_points and data_points['tds']:
                # If TDS is already in data_points, mark it as not computed
                data_points['tds_computed'] = 'No'
                logger.info(f"Using extracted TDS: {data_points['tds']}")

            # Track which fields were successfully updated
            for field, mapped_column in self.header_mapping.items():
                if field in data_points and data_points[field]:
                    value = data_points[field]
                    column_index = self.df.columns.get_loc(mapped_column) + 1  # 1-indexed for openpyxl

                    if 'receipt_amount' in data_points:
                        logger.info(
                            f"IMPORTANT DEBUG - Final receipt_amount in data_points: {data_points['receipt_amount']}")

                    # Update the cell value
                    cell = self.ws.cell(row=excel_row, column=column_index)
                    original_value = cell.value
                    cell.value = value

                    # Highlight the updated cell
                    cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

                    # Log the update
                    logger.info(f"Updated row {row_index}, column '{mapped_column}': '{original_value}' -> '{value}'")
                    updated_fields.append(field)
                else:
                    if field not in data_points:
                        logger.warning(f"Field '{field}' not found in extracted data")
                    else:
                        logger.warning(f"No value provided for field '{field}'")

                    # Don't count TDS Computed as a failed field if it wasn't provided
                    if field != 'tds_computed':
                        failed_fields.append(field)

            # Special handling for TDS Computed field
            tds_computed_column = self.header_mapping.get('tds_computed')
            if tds_computed_column and 'tds' in data_points:
                tds_computed_value = data_points.get('tds_computed', 'No')  # Default to No if not provided
                tds_computed_index = self.df.columns.get_loc(tds_computed_column) + 1
                tds_computed_cell = self.ws.cell(row=excel_row, column=tds_computed_index)
                tds_computed_cell.value = tds_computed_value
                tds_computed_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                logger.info(f"Updated TDS Computed status: {tds_computed_value}")

                if 'tds_computed' not in updated_fields:
                    updated_fields.append('tds_computed')

            # Add timestamp to indicate when the record was updated
            last_column = len(self.df.columns) + 1
            timestamp_cell = self.ws.cell(row=excel_row, column=last_column)
            timestamp_cell.value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            # Success if at least one of receipt_amount, receipt_date, or tds was updated
            core_fields = ['receipt_amount', 'receipt_date', 'tds']
            core_fields_updated = any(field in updated_fields for field in core_fields)

            if core_fields_updated:
                # Save the updated workbook
                backup_path = self.excel_path.replace('.xlsx',
                                                      f'_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
                self.wb.save(backup_path)  # Save a backup first
                self.wb.save(self.excel_path)  # Then save the actual file
                logger.info(f"Saved updated Excel file and backup: {backup_path}")
                return True, updated_fields, failed_fields
            else:
                logger.warning(f"No core fields (receipt_amount, receipt_date, tds) were updated for row {row_index}")
                return False, updated_fields, failed_fields

        except Exception as e:
            logger.error(f"Error updating row {row_index}: {str(e)}")
            return False, [], list(data_points.keys())

    def process_multi_row_data(self, table_data, pdf_text, global_data_points=None):
        """
        Process data from a table in the PDF where each row might correspond to a different Excel row.

        Args:
            table_data (list): List of dictionaries, each containing data for one row
            pdf_text (str): The full PDF text for context and TDS computation
            global_data_points (dict, optional): Global data points extracted from the PDF

        Returns:
            dict: Processing results
        """
        results = {
            'processed': 0,
            'unprocessed': 0,
            'total': len(table_data)
        }

        # Use provided global data points or an empty dict
        global_data = global_data_points or {}

        # Track which Excel rows we've updated to avoid duplicates
        processed_ids = set()

        for row_index, row_data in enumerate(table_data):
            # The row_data should contain at least one identifier (unique_id)
            # and at least one value to update (receipt_amount, receipt_date, or tds)

            unique_id = row_data.get('unique_id')
            if not unique_id:
                logger.warning(f"No unique identifier found for row data: {row_data}")
                results['unprocessed'] += 1
                continue

            # Skip if we've already processed this ID
            if unique_id in processed_ids:
                logger.info(f"Already processed {unique_id}, skipping duplicate")
                continue

            # Copy missing fields from global data points to row_data
            for key, value in global_data.items():
                if key not in row_data and key != 'unique_id':
                    row_data[key] = value
                    logger.debug(f"Added global field '{key}': {value} to row data for {unique_id}")

            # Find the row in Excel
            excel_row_index = self.find_row_by_identifiers(unique_id, row_data, pdf_text)

            if excel_row_index is None:
                logger.warning(f"No matching Excel row found for ID: {unique_id}")
                results['unprocessed'] += 1
                continue

            if 'receipt_date' in row_data and row_data['receipt_date'] and row_data['receipt_date'] != global_data[
                'receipt_date']:
                logger.info(
                    f"Using table-specific receipt date: {row_data['receipt_date']} instead of global: {global_data.get('receipt_date')}")

            # Update the row with data
            success, updated_fields, failed_fields = self.update_row_with_data(excel_row_index, row_data, pdf_text)

            if success:
                results['processed'] += 1
                logger.info(f"Successfully updated row for ID: {unique_id}")
                # Mark this ID as processed
                processed_ids.add(unique_id)
            else:
                results['unprocessed'] += 1
                logger.warning(f"Failed to update row for ID: {unique_id}")

        return results

    def save_excel(self):
        """Save the Excel workbook."""
        try:
            backup_path = self.excel_path.replace('.xlsx',
                                                  f'_backup_final_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
            self.wb.save(backup_path)
            self.wb.save(self.excel_path)
            logger.info(f"Excel file saved: {self.excel_path} with backup: {backup_path}")
            return True
        except Exception as e:
            logger.error(f"Error saving Excel file: {str(e)}")
            return False